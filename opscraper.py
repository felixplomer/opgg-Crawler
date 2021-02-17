
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlsxwriter
from tkinter import *

'''
def excel_sheet(ranks, player):
    # übertrage Ergebnisse in BESTEHENDES Excel Sheet
    wb = load_workbook("D:\Python_Projects\Edabit_Challenges\AnalyseSheet.xlsx")
    sheets = wb.sheetnames
    Sheet1 = wb[sheets[0]]

    Sheet1.cell(row=1, column=1).value = player[0]
    Sheet1.cell(row=2, column=1).value = ranks[0]
    Sheet1.cell(row=1, column=2).value = player[1]
    Sheet1.cell(row=2, column=2).value = ranks[1]
    Sheet1.cell(row=1, column=3).value = player[2]
    Sheet1.cell(row=2, column=3).value = ranks[2]
    Sheet1.cell(row=1, column=4).value = player[3]
    Sheet1.cell(row=2, column=4).value = ranks[3]
    Sheet1.cell(row=1, column=5).value = player[4]
    Sheet1.cell(row=2, column=5).value = ranks[4]

    wb.save("D:\Python_Projects\Edabit_Challenges\AnalyseSheet.xlsx")
'''


def excel_sheet(ranks,player, most_played_champs):

    # übertrage Ergebnisse in NEUES Excel Sheet

    workbook = xlsxwriter.Workbook("AnalyseSheet.xlsx")
    worksheet = workbook.add_worksheet()
    worksheet.set_column("A:A", 20)
    worksheet.set_column("B:F", 15)

    bold = workbook.add_format({"bold" : True})


    worksheet.write("A1", "Player Name", bold)
    worksheet.write("B1", player[0])
    worksheet.write("C1", player[1])
    worksheet.write("D1", player[2])
    worksheet.write("E1", player[3])
    worksheet.write("F1", player[4])

    worksheet.write("A2", "Player Solo Rank", bold)
    worksheet.write("B2", ranks[0])
    worksheet.write("C2", ranks[1])
    worksheet.write("D2", ranks[2])
    worksheet.write("E2", ranks[3])
    worksheet.write("F2", ranks[4])

    worksheet.write("A3", "Most Played Champs", bold)
    worksheet.write("B3", most_played_champs[0][0])
    worksheet.write("B4", most_played_champs[0][1])
    worksheet.write("B5", most_played_champs[0][2])

    worksheet.write("C3", most_played_champs[1][0])
    worksheet.write("C4", most_played_champs[1][1])
    worksheet.write("C5", most_played_champs[1][2])

    worksheet.write("D3", most_played_champs[2][0])
    worksheet.write("D4", most_played_champs[2][1])
    worksheet.write("D5", most_played_champs[2][2])

    worksheet.write("E3", most_played_champs[3][0])
    worksheet.write("E4", most_played_champs[3][1])
    worksheet.write("E5", most_played_champs[3][2])

    worksheet.write("F3", most_played_champs[4][0])
    worksheet.write("F4", most_played_champs[4][1])
    worksheet.write("F5", most_played_champs[4][2])


    workbook.close()



def gui():
    UserName = spieler1.get()
    UserName2 = spieler2.get()
    UserName3 = spieler3.get()
    UserName4 = spieler4.get()
    UserName5 = spieler5.get()

    player = [UserName, UserName2, UserName3, UserName4, UserName5]

    master.quit()
    return player


master = Tk()
master.title("Analyse Tool")
master.geometry("300x200")


Label(master, text="Spieler 1").grid(row=0)
Label(master, text="Spieler 2").grid(row=1)
Label(master, text="Spieler 3").grid(row=2)
Label(master, text="Spieler 4").grid(row=3)
Label(master, text="Spieler 5").grid(row=4)

spieler1 = Entry(master)
spieler2 = Entry(master)
spieler3 = Entry(master)
spieler4 = Entry(master)
spieler5 = Entry(master)

spieler1.grid(row=0, column=2)
spieler2.grid(row=1, column=2)
spieler3.grid(row=2, column=2)
spieler4.grid(row=3, column=2)
spieler5.grid(row=4, column=2)


Button(master, text='Confirm', command=gui,).grid(row=6, column=2, sticky=W, pady=4)
Button(master, text='Details', command= master.quit,).grid(row=6, column=3, sticky=W, pady=5)

mainloop()



def rang_update(ranks, player):
    master = Tk()
    master.geometry("300x200")


    Label(master, text= player[0] + " : " + ranks[0]).grid(row=0, column = 3)
    Label(master, text= player[1] + " : " + ranks[1]).grid(row=1, column = 3)
    Label(master, text= player[2] + " : " + ranks[2]).grid(row=2, column = 3)
    Label(master, text= player[3] + " : " + ranks[3]).grid(row=3, column = 3)
    Label(master, text= player[4] + " : " + ranks[4]).grid(row=4, column = 3)

    mainloop()



ranks = []
most_played_champs = []

for user in gui():


    URL = "https://euw.op.gg/summoner/userName=" + user
    page = requests.get(URL)

    soup = BeautifulSoup(page.content, "html.parser")
    # player info
    solo_rank = soup.find("div", class_="TierRank")
    #flex_rank = soup.find("div", class_="sub-tier__rank-tier")

    if solo_rank.text == "\n\t\t\tUnranked\n\t\t":
        solo_rank = soup.find_all("li", class_ = "Item tip")
        for entry in solo_rank:
            if entry.find("b", string="S2020"):
                solo_rank = entry
    ranks.append(solo_rank.text)

    # Most Played Champions
    most_played = []

    URL = "https://euw.op.gg/summoner/champions/userName=" + user
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, "html.parser")

    champions_played_most = soup.find_all("tr", class_="Row TopRanker")
    for i, x in enumerate(champions_played_most):
        if i < 3:
            champion = x.find("td", class_="ChampionName Cell")
            most_played.append(champion.text)

    most_played_champs.append(most_played)

print(most_played_champs)



excel_sheet(ranks, gui(), most_played_champs)















'''
# last 10 games history stats und farm

game_history = soup.find("div", class_="GameItemList")
games = game_history.find_all("div", class_="GameItemWrap")

for game in games:
    champion = game.find("div", class_="ChampionName")

    kills = game.find("span", class_="Kill")
    death = game.find("span", class_="Death")
    assists = game.find("span", class_="Assist")
    farm = game.find("span", class_="CS")

    print(champion.text)
    print("CS: ", farm.text)
    print(kills.text, "/", death.text, "/", assists.text)

'''



